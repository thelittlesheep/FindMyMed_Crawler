import openpyxl
import os

'''def top100(sheet, drug107_list, drug106_list, drug105_list):
    for i in range(2, 105):
        #2~105，d1^d2^d3 = top100drugs
        drug107_list.append(sheet.cell(row=i, column=3).value)
        drug106_list.append(sheet.cell(row=i, column=6).value)
        drug105_list.append(sheet.cell(row=i, column=9).value)

    d1 = set(drug107_list)
    d2 = drug106_list
    d3 = drug105_list
    d4 = [i for i in d2 if i in d1]
    d5 = [i for i in d3 if i in d4]
    top100drugsname = []
    for i in range(0, len(d5)):
        top100drugsname.append(list(d5)[i].split(',')[0].split(' ')[0])

    return top100drugsname'''


def all_drug(sheet, drug107_list, drug106_list, drug105_list):
    for i in range(2, 3382):
        drug107_list.append(sheet.cell(row=i, column=3).value)
        drug106_list.append(sheet.cell(row=i, column=6).value)
        drug105_list.append(sheet.cell(row=i, column=9).value)

    d1 = set(drug107_list)
    d2 = drug106_list
    d3 = drug105_list
    d4 = [i for i in d2 if i in d1]
    d5 = [i for i in d3 if i in d4]
    all_drug = []
    all_drug_raw = []
    for i in range(0, len(d5)):
        all_drug.append(list(d5)[i].split(',')[0].split(' ')[0])
    for i in range(0, len(d5)):
        all_drug_raw.append(str(i) + "." + list(d5)[i])
    return all_drug


def main():
    # os.chdir 是 python 切換到電腦指定路徑的方法
    # os.chdir(r"C:\Users\ejiej\OneDrive\作業與上課講義\專題\用藥資料")
    workbook = openpyxl.load_workbook('107~105.xlsx')
    sheet = workbook['book1']

    drug107_list = []
    drug106_list = []
    drug105_list = []
    a = all_drug(sheet, drug107_list, drug106_list, drug105_list)
    # b = top100(sheet, drug107_list, drug106_list, drug105_list)
    # print(len(a))
    # print(a)

    return a


if __name__ == '__main__':
    main()
